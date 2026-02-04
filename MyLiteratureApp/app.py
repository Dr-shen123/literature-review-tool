# literature_reviewer_with_custom_columns_v2.py
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import os
import tempfile
import warnings
from datetime import datetime

# 忽略警告
warnings.filterwarnings('ignore')

# ====================== 页面配置 ======================
st.set_page_config(
    page_title="文献筛选工具",
    layout="wide",
    initial_sidebar_state="expanded",
    page_icon="📚"
)

# ====================== 基础样式 ======================
base_css = """
<style>
/* 主容器 */
.main-container {
    max-width: 1200px;
    margin: 0 auto;
}

/* 文献卡片 */
.paper-card {
    background: white;
    border-radius: 8px;
    padding: 24px;
    margin: 20px 0;
    box-shadow: 0 2px 10px rgba(0,0,0,0.05);
    border-left: 4px solid #4285f4;
}

/* 内容区域 */
.content-section {
    margin-bottom: 20px;
    line-height: 1.6;
}

.content-section h4 {
    color: #1a73e8;
    margin-bottom: 8px;
    border-bottom: 1px solid #eee;
    padding-bottom: 4px;
}

/* 分类按钮 */
.classification-btn {
    padding: 12px 24px;
    font-size: 16px;
    font-weight: bold;
    border-radius: 6px;
    border: none;
    cursor: pointer;
    transition: all 0.3s ease;
    margin: 0 5px;
}

.classification-btn:hover {
    transform: translateY(-2px);
    box-shadow: 0 4px 8px rgba(0,0,0,0.1);
}

.include-btn {
    background-color: #4CAF50;
    color: white;
}

.exclude-btn {
    background-color: #f44336;
    color: white;
}

.pending-btn {
    background-color: #ff9800;
    color: white;
}

/* 状态标签 */
.status-badge {
    display: inline-block;
    padding: 4px 12px;
    border-radius: 12px;
    font-size: 12px;
    font-weight: bold;
    margin-left: 10px;
}

.status-include { background-color: #e8f5e9; color: #2e7d32; }
.status-exclude { background-color: #ffebee; color: #c62828; }
.status-pending { background-color: #fff3e0; color: #ef6c00; }

/* 备注区域 */
.note-section {
    background-color: #f8f9fa;
    border-radius: 6px;
    padding: 16px;
    margin: 20px 0;
}

/* 导航按钮 */
.nav-btn {
    margin: 0 5px;
    min-width: 100px;
}

/* 字体大小预览 */
.font-preview {
    margin: 10px 0;
    padding: 10px;
    background-color: #f8f9fa;
    border-radius: 4px;
    border: 1px dashed #ddd;
}

/* 自定义列区域 */
.custom-columns-section {
    background-color: #f0f7ff;
    border-radius: 6px;
    padding: 15px;
    margin: 15px 0;
    border: 1px solid #d0e3ff;
}

/* 自定义列配置项 */
.column-config-item {
    background-color: #f9f9f9;
    border-radius: 4px;
    padding: 10px;
    margin-bottom: 10px;
    border-left: 4px solid #4CAF50;
}
</style>
"""

st.markdown(base_css, unsafe_allow_html=True)

# ====================== 初始化Session State ======================
def initialize_session_state():
    """初始化所有session state变量"""
    defaults = {
        'df': None,
        'column_mapping': {},
        'current_index': 0,
        'selections': {},
        'notes': {},
        'file_processed': False,
        'current_filename': None,
        'show_column_mapping': False,
        'mapping_confirmed': False,
        'auto_advance': True,
        'font_size': 16,  # 默认字体大小
        'font_size_abstract': 14,  # 摘要字体大小
        'font_size_translation': 14,  # 翻译字体大小
        'extra_columns': {},  # 存储自定义列配置 {列名: {display_name, position, collapsed}}
        'show_extra_columns': True  # 是否显示自定义列
    }
    
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

# ====================== 工具函数 ======================
def detect_column_candidates(df):
    """检测可能的列名候选"""
    columns = df.columns.tolist()
    candidates = {
        'title': [],
        'title_translation': [],
        'abstract': [],
        'abstract_translation': []
    }
    
    # 常见列名关键词
    title_keywords = ['标题', 'title', '题名', '篇名', '文章标题', '题目', 'ti']
    translation_keywords = ['翻译', 'translation', '英文', 'english', 'en']
    abstract_keywords = ['摘要', 'abstract', '概要', '内容简介', '文章摘要', 'ab']
    
    for col in columns:
        col_lower = str(col).lower()
        
        # 检查标题
        if any(keyword in col_lower for keyword in title_keywords):
            if any(keyword in col_lower for keyword in translation_keywords):
                candidates['title_translation'].append(col)
            else:
                candidates['title'].append(col)
        
        # 检查摘要
        elif any(keyword in col_lower for keyword in abstract_keywords):
            if any(keyword in col_lower for keyword in translation_keywords):
                candidates['abstract_translation'].append(col)
            else:
                candidates['abstract'].append(col)
    
    return candidates

def save_results():
    """保存处理结果到Excel（包含四个工作表）"""
    if st.session_state.df is None:
        st.error("没有数据可保存")
        return None
    
    df = st.session_state.df
    column_mapping = st.session_state.column_mapping
    
    # 创建结果DataFrame（主工作表）
    result_df = df.copy()
    
    # 确保有备注列
    if '备注' not in result_df.columns:
        result_df['备注'] = ''
    
    # 更新备注
    for i in range(len(result_df)):
        note_key = f"note_{i}"
        if note_key in st.session_state.notes:
            result_df.at[i, '备注'] = st.session_state.notes[note_key]
    
    # 为每个分类创建DataFrame
    include_indices = []
    pending_indices = []
    exclude_indices = []
    
    # 获取每个分类的索引
    for idx in range(len(df)):
        if idx in st.session_state.selections:
            selection = st.session_state.selections[idx]
            if selection == '纳入':
                include_indices.append(idx)
            elif selection == '待定':
                pending_indices.append(idx)
            elif selection == '排除':
                exclude_indices.append(idx)
    
    # 创建分类DataFrame
    df_include = df.iloc[include_indices].copy() if include_indices else pd.DataFrame(columns=df.columns)
    df_pending = df.iloc[pending_indices].copy() if pending_indices else pd.DataFrame(columns=df.columns)
    df_exclude = df.iloc[exclude_indices].copy() if exclude_indices else pd.DataFrame(columns=df.columns)
    
    # 为分类DataFrame添加备注
    for idx in include_indices:
        note_key = f"note_{idx}"
        if note_key in st.session_state.notes:
            df_include.at[idx, '备注'] = st.session_state.notes[note_key]
    
    for idx in pending_indices:
        note_key = f"note_{idx}"
        if note_key in st.session_state.notes:
            df_pending.at[idx, '备注'] = st.session_state.notes[note_key]
    
    for idx in exclude_indices:
        note_key = f"note_{idx}"
        if note_key in st.session_state.notes:
            df_exclude.at[idx, '备注'] = st.session_state.notes[note_key]
    
    # 保存到临时文件
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            temp_path = tmp_file.name
            
            # 使用pandas的ExcelWriter写入多个工作表
            with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
                # 写入主工作表（所有文献）
                result_df.to_excel(writer, sheet_name='所有文献', index=False)
                
                # 写入分类工作表
                df_include.to_excel(writer, sheet_name='纳入文章', index=False)
                df_pending.to_excel(writer, sheet_name='待定文章', index=False)
                df_exclude.to_excel(writer, sheet_name='排除文章', index=False)
            
            # 重新打开工作簿设置颜色格式
            wb = openpyxl.load_workbook(temp_path)
            
            # 为"所有文献"工作表设置颜色标记
            ws_all = wb['所有文献']
            
            # 颜色填充定义
            red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
            green_fill = PatternFill(start_color='FF90EE90', end_color='FF90EE90', fill_type='solid')  # 添加绿色填充
            
            # 从第二行开始（第一行是标题）
            for i, row in enumerate(ws_all.iter_rows(min_row=2, max_row=len(df)+1), start=0):
                if i in st.session_state.selections:
                    selection = st.session_state.selections[i]
                    cell = row[0]  # 第一列（序号列）
                    
                    if selection == '排除':
                        cell.fill = red_fill
                    elif selection == '待定':
                        cell.fill = yellow_fill
                    elif selection == '纳入':  # 添加对纳入文献的处理
                        cell.fill = green_fill
            
            # 为分类工作表的序号列添加简单格式（可选）
            for sheet_name, df_sheet in [('纳入文章', df_include), ('待定文章', df_pending), ('排除文章', df_exclude)]:
                if sheet_name in wb.sheetnames and len(df_sheet) > 0:
                    ws_sheet = wb[sheet_name]
                    # 可以根据需要为分类工作表的序号列添加不同颜色
                    # 例如：纳入文章用绿色，待定文章用橙色，排除文章用红色
                    if sheet_name == '纳入文章':
                        fill_color = PatternFill(start_color='FF90EE90', end_color='FF90EE90', fill_type='solid')
                    elif sheet_name == '待定文章':
                        fill_color = PatternFill(start_color='FFFFE0B2', end_color='FFFFE0B2', fill_type='solid')
                    elif sheet_name == '排除文章':
                        fill_color = PatternFill(start_color='FFFFCCCC', end_color='FFFFCCCC', fill_type='solid')
                    
                    for i, row in enumerate(ws_sheet.iter_rows(min_row=2, max_row=len(df_sheet)+1), start=1):
                        cell = row[0]  # 第一列（序号列）
                        cell.fill = fill_color
            
            wb.save(temp_path)
        
        return temp_path
        
    except Exception as e:
        st.error(f"保存文件时出错: {str(e)}")
        return None

def handle_classification(selection):
    """处理分类选择"""
    df = st.session_state.df
    current_idx = st.session_state.current_index
    
    # 保存当前备注
    note_key = f"note_{current_idx}"
    if 'current_note' in st.session_state:
        st.session_state.notes[note_key] = st.session_state.current_note
    
    # 记录分类选择
    st.session_state.selections[current_idx] = selection
    
    # 检查是否启用自动跳转
    if st.session_state.auto_advance and current_idx < len(df) - 1:
        st.session_state.current_index += 1
        st.rerun()

def display_custom_column_value(value, col_name, current_idx):
    """显示自定义列的值"""
    if pd.isna(value):
        return ""
    
    value_str = str(value)
    # 根据内容长度决定显示方式
    if len(value_str) > 200:
        return st.text_area("", value=value_str, height=100, 
                          key=f"extra_{col_name}_{current_idx}", disabled=True, label_visibility="collapsed")
    else:
        return st.markdown(f'<div style="padding: 8px; background-color: #f8f9fa; border-radius: 4px; margin-bottom: 10px;">{value_str}</div>', 
                         unsafe_allow_html=True)

# ====================== 字体大小设置界面 ======================
def create_font_settings_ui():
    """创建字体大小设置界面"""
    with st.expander("🎨 字体大小设置", expanded=False):
        col1, col2 = st.columns(2)
        
        with col1:
            # 摘要字体大小
            font_size_abstract = st.slider(
                "摘要字体大小",
                min_value=10,
                max_value=24,
                value=st.session_state.font_size_abstract,
                step=1,
                help="调整摘要的显示字体大小"
            )
            st.session_state.font_size_abstract = font_size_abstract
            
            # 预览
            st.markdown(f'<div class="font-preview" style="font-size: {font_size_abstract}px;">摘要预览：这是一个示例文本，使用当前字体大小显示。</div>', 
                       unsafe_allow_html=True)
        
        with col2:
            # 翻译字体大小
            font_size_translation = st.slider(
                "翻译字体大小",
                min_value=10,
                max_value=24,
                value=st.session_state.font_size_translation,
                step=1,
                help="调整翻译内容的显示字体大小"
            )
            st.session_state.font_size_translation = font_size_translation
            
            # 预览
            st.markdown(f'<div class="font-preview" style="font-size: {font_size_translation}px;">翻译预览：This is a sample text showing current font size.</div>', 
                       unsafe_allow_html=True)
        
        # 重置按钮
        if st.button("重置为默认大小", use_container_width=True):
            st.session_state.font_size_abstract = 14
            st.session_state.font_size_translation = 14
            st.success("字体大小已重置")
            st.rerun()

# ====================== 主应用 ======================
def main():
    # 初始化session state
    initialize_session_state()
    
    # 应用标题
    st.title("📚 文献筛选工具")
    st.markdown("---")
    
    # ====================== 左侧边栏 ======================
    with st.sidebar:
        st.header("📁 文件管理")
        
        # 文件上传
        uploaded_file = st.file_uploader(
            "上传Excel文件",
            type=['xlsx', 'xls'],
            help="请上传包含文献信息的Excel文件"
        )
        
        if uploaded_file:
            if not st.session_state.file_processed or uploaded_file.name != st.session_state.current_filename:
                try:
                    df = pd.read_excel(uploaded_file)
                    
                    # 确保有序号列
                    if '序号' not in df.columns:
                        df.insert(0, '序号', range(1, len(df) + 1))
                    
                    st.session_state.df = df
                    st.session_state.current_filename = uploaded_file.name
                    st.session_state.file_processed = True
                    st.session_state.show_column_mapping = True
                    st.session_state.mapping_confirmed = False
                    st.session_state.current_index = 0
                    st.session_state.selections = {}
                    st.session_state.notes = {}
                    st.session_state.extra_columns = {}  # 重置自定义列配置
                    
                    st.success(f"成功加载 {len(df)} 篇文献")
                    
                except Exception as e:
                    st.error(f"读取文件失败: {str(e)}")
        
        # 字体大小设置（全局显示）
        create_font_settings_ui()
        
        # 列映射配置
        if st.session_state.df is not None and not st.session_state.mapping_confirmed:
            st.header("🔧 列映射配置")
            
            df = st.session_state.df
            columns = [""] + df.columns.tolist()
            candidates = detect_column_candidates(df)
            
            # 标题列选择
            title_default = candidates['title'][0] if candidates['title'] else ""
            title_col = st.selectbox(
                "选择标题列",
                options=columns,
                index=columns.index(title_default) if title_default in columns else 0,
                key="title_select"
            )
            
            # 标题翻译列选择
            title_trans_default = candidates['title_translation'][0] if candidates['title_translation'] else ""
            title_trans_col = st.selectbox(
                "选择标题翻译列（可选）",
                options=columns,
                index=columns.index(title_trans_default) if title_trans_default in columns else 0,
                key="title_trans_select"
            )
            
            # 摘要列选择
            abstract_default = candidates['abstract'][0] if candidates['abstract'] else ""
            abstract_col = st.selectbox(
                "选择摘要列",
                options=columns,
                index=columns.index(abstract_default) if abstract_default in columns else 0,
                key="abstract_select"
            )
            
            # 摘要翻译列选择
            abstract_trans_default = candidates['abstract_translation'][0] if candidates['abstract_translation'] else ""
            abstract_trans_col = st.selectbox(
                "选择摘要翻译列（可选）",
                options=columns,
                index=columns.index(abstract_trans_default) if abstract_trans_default in columns else 0,
                key="abstract_trans_select"
            )
            
            # 自定义列配置
            st.subheader("🔍 自定义显示列配置")
            st.markdown('<div class="custom-columns-section">', unsafe_allow_html=True)
            
            # 选择要显示的额外列
            available_columns = [col for col in df.columns if col not in [title_col, title_trans_col, abstract_col, abstract_trans_col, '序号', '备注']]
            
            if available_columns:
                # 初始化extra_columns
                if 'extra_columns' not in st.session_state:
                    st.session_state.extra_columns = {}
                
                extra_cols_selected = st.multiselect(
                    "选择要显示的额外列",
                    options=available_columns,
                    default=list(st.session_state.extra_columns.keys()),
                    help="选择需要在文献详情中显示的额外列"
                )
                
                # 为每个选中的列配置显示选项
                st.markdown("**列显示配置**")
                
                # 获取现有的配置或创建新的
                extra_cols_config = st.session_state.extra_columns.copy()
                
                # 为每个选中的列创建配置项
                for i, col in enumerate(extra_cols_selected):
                    st.markdown(f'<div class="column-config-item">', unsafe_allow_html=True)
                    st.markdown(f"**列{i+1}: `{col}`**")
                    
                    # 获取现有配置或创建默认配置
                    if col in extra_cols_config:
                        col_config = extra_cols_config[col]
                    else:
                        col_config = {
                            'display_name': col,
                            'position': '原文信息栏',
                            'collapsed': True
                        }
                    
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        # 显示名称
                        display_name = st.text_input(
                            "显示名称",
                            value=col_config['display_name'],
                            key=f"display_name_{col}",
                            help="输入该列的显示名称"
                        )
                    
                    with col2:
                        # 位置选择
                        position = st.selectbox(
                            "显示位置",
                            options=['原文信息栏', '翻译信息栏', '分类选择后'],
                            index=['原文信息栏', '翻译信息栏', '分类选择后'].index(col_config['position']),
                            key=f"position_{col}",
                            help="选择该列在页面中的显示位置"
                        )
                    
                    with col3:
                        # 是否折叠
                        collapsed = st.checkbox(
                            "折叠显示",
                            value=col_config['collapsed'],
                            key=f"collapsed_{col}",
                            help="勾选后该列将在折叠区域中显示"
                        )
                    
                    # 更新配置
                    extra_cols_config[col] = {
                        'display_name': display_name,
                        'position': position,
                        'collapsed': collapsed
                    }
                    st.markdown('</div>', unsafe_allow_html=True)
                
                # 移除未选中的列的配置
                cols_to_remove = [col for col in extra_cols_config if col not in extra_cols_selected]
                for col in cols_to_remove:
                    del extra_cols_config[col]
                
                st.session_state.extra_columns = extra_cols_config
            else:
                st.info("没有可用的额外列")
                st.session_state.extra_columns = {}
            
            st.markdown('</div>', unsafe_allow_html=True)
            
            col1, col2 = st.columns(2)
            
            with col1:
                if st.button("确认映射", type="primary", use_container_width=True):
                    if not title_col or not abstract_col:
                        st.error("请至少选择标题列和摘要列")
                    else:
                        st.session_state.column_mapping = {
                            'title': title_col,
                            'title_translation': title_trans_col if title_trans_col else None,
                            'abstract': abstract_col,
                            'abstract_translation': abstract_trans_col if abstract_trans_col else None
                        }
                        st.session_state.mapping_confirmed = True
                        st.success("列映射已确认！")
                        st.rerun()
            
            with col2:
                if st.button("重置", type="secondary", use_container_width=True):
                    st.session_state.column_mapping = {}
                    st.session_state.extra_columns = {}
                    st.rerun()
        
        # 导航与设置（如果映射已确认）
        if st.session_state.df is not None and st.session_state.mapping_confirmed:
            df = st.session_state.df
            
            st.header("⚙️ 设置与导航")
            
            # 自动跳转设置
            st.session_state.auto_advance = st.checkbox(
                "选择分类后自动跳转到下一篇",
                value=st.session_state.auto_advance,
                help="启用后，选择分类会自动保存并显示下一篇文献"
            )
            
            current_idx = st.session_state.current_index
            
            # 导航控制
            col_nav1, col_nav2 = st.columns(2)
            with col_nav1:
                if st.button("◀ 上一篇", disabled=current_idx <= 0, use_container_width=True):
                    st.session_state.current_index -= 1
                    st.rerun()
            
            with col_nav2:
                if st.button("下一篇 ▶", disabled=current_idx >= len(df)-1, use_container_width=True):
                    st.session_state.current_index += 1
                    st.rerun()
            
            # 快速跳转
            target_idx = st.number_input(
                "跳转到文献序号",
                min_value=1,
                max_value=len(df),
                value=current_idx + 1,
                key="jump_input"
            )
            
            if target_idx - 1 != current_idx:
                st.session_state.current_index = target_idx - 1
                st.rerun()
            
            # 进度统计
            st.header("📊 进度统计")
            
            total = len(df)
            processed = len(st.session_state.selections)
            progress = processed / total if total > 0 else 0
            
            st.progress(progress)
            st.write(f"**已处理**: {processed}/{total} 篇 ({progress:.1%})")
            
            # 分类统计
            if st.session_state.selections:
                from collections import Counter
                counts = Counter(st.session_state.selections.values())
                
                col_stat1, col_stat2, col_stat3 = st.columns(3)
                with col_stat1:
                    st.metric("纳入", counts.get('纳入', 0))
                with col_stat2:
                    st.metric("排除", counts.get('排除', 0))
                with col_stat3:
                    st.metric("待定", counts.get('待定', 0))
            
            # 保存导出
            st.header("💾 保存导出")
            
            # 显示导出说明
            st.info("导出将生成包含以下工作表的Excel文件：\n1. 所有文献（带颜色标记）\n2. 纳入文章\n3. 待定文章\n4. 排除文章")
            
            if st.button("保存进度并导出", type="primary", use_container_width=True):
                temp_path = save_results()
                
                if temp_path:
                    # 提供下载
                    with open(temp_path, 'rb') as f:
                        st.download_button(
                            label="📥 下载Excel文件",
                            data=f,
                            file_name=f"文献筛选结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    
                    # 清理临时文件
                    os.unlink(temp_path)
    
    # ====================== 主内容区域 ======================
    if st.session_state.df is not None and st.session_state.mapping_confirmed:
        df = st.session_state.df
        current_idx = st.session_state.current_index
        column_mapping = st.session_state.column_mapping
        
        # 创建文献卡片
        st.markdown('<div class="paper-card">', unsafe_allow_html=True)
        
        # 顶部状态栏
        col_top1, col_top2 = st.columns([4, 1])
        
        with col_top1:
            st.markdown(f"### 文献 #{current_idx + 1}")
        
        with col_top2:
            if current_idx in st.session_state.selections:
                status = st.session_state.selections[current_idx]
                status_class = f"status-{status}"
                st.markdown(f'<div class="status-badge {status_class}">{status}</div>', unsafe_allow_html=True)
        
        st.markdown("---")
        
        # 双栏显示文献内容
        col_content1, col_content2 = st.columns(2)
        
        with col_content1:
            st.markdown('<div class="content-section">', unsafe_allow_html=True)
            st.markdown("#### 原文信息")
            
            # 标题
            title_col = column_mapping.get('title')
            if title_col and title_col in df.columns:
                title = df.iloc[current_idx][title_col]
                if pd.notna(title):
                    st.markdown("**标题**")
                    st.markdown(f'<div style="margin-bottom: 15px; padding: 10px; background-color: #f8f9fa; border-radius: 4px; font-size: 18px;">{title}</div>', 
                               unsafe_allow_html=True)
            
            # 摘要
            abstract_col = column_mapping.get('abstract')
            if abstract_col and abstract_col in df.columns:
                abstract = df.iloc[current_idx][abstract_col]
                if pd.notna(abstract):
                    st.markdown("**摘要**")
                    # 使用动态字体大小
                    font_size = st.session_state.font_size_abstract
                    st.markdown(f'<div style="white-space: pre-wrap; line-height: 1.6; margin-bottom: 20px; font-size: {font_size}px;">{abstract}</div>', 
                               unsafe_allow_html=True)
            
            # 显示位置在"原文信息栏"的自定义列
            display_custom_columns_by_position('原文信息栏', df, current_idx)
            
            st.markdown('</div>', unsafe_allow_html=True)
        
        with col_content2:
            st.markdown('<div class="content-section">', unsafe_allow_html=True)
            st.markdown("#### 翻译信息")
            
            # 标题翻译
            title_trans_col = column_mapping.get('title_translation')
            if title_trans_col and title_trans_col in df.columns:
                title_trans = df.iloc[current_idx][title_trans_col]
                if pd.notna(title_trans):
                    st.markdown("**标题翻译**")
                    st.markdown(f'<div style="margin-bottom: 15px; padding: 10px; background-color: #e8f5e9; border-radius: 4px; font-size: 18px;">{title_trans}</div>', 
                               unsafe_allow_html=True)
            else:
                st.info("无标题翻译信息")
            
            # 摘要翻译
            abstract_trans_col = column_mapping.get('abstract_translation')
            if abstract_trans_col and abstract_trans_col in df.columns:
                abstract_trans = df.iloc[current_idx][abstract_trans_col]
                if pd.notna(abstract_trans):
                    st.markdown("**摘要翻译**")
                    # 使用动态字体大小
                    font_size = st.session_state.font_size_translation
                    st.markdown(f'<div style="white-space: pre-wrap; line-height: 1.6; margin-bottom: 20px; font-size: {font_size}px;">{abstract_trans}</div>', 
                               unsafe_allow_html=True)
            
            # 显示位置在"翻译信息栏"的自定义列
            display_custom_columns_by_position('翻译信息栏', df, current_idx)
            
            st.markdown('</div>', unsafe_allow_html=True)
        
        st.markdown('</div>', unsafe_allow_html=True)
        
        # 分类按钮区域
        st.markdown("### 🏷️ 分类选择")
        
        col_btn1, col_btn2, col_btn3, col_btn4 = st.columns(4)
        
        with col_btn1:
            if st.button("✅ 纳入", key="include_btn", use_container_width=True):
                handle_classification('纳入')
        
        with col_btn2:
            if st.button("❌ 排除", key="exclude_btn", use_container_width=True):
                handle_classification('排除')
        
        with col_btn3:
            if st.button("⚠️ 待定", key="pending_btn", use_container_width=True):
                handle_classification('待定')
        
        with col_btn4:
            if st.button("⏸️ 暂停跳转", key="pause_btn", use_container_width=True, type="secondary"):
                st.session_state.auto_advance = not st.session_state.auto_advance
                status = "已启用" if st.session_state.auto_advance else "已暂停"
                st.success(f"自动跳转{status}")
                st.rerun()
        
        # 显示自动跳转状态
        if st.session_state.auto_advance:
            st.info("自动跳转已启用 - 选择分类后将自动跳转到下一篇")
        else:
            st.warning("自动跳转已暂停 - 选择分类后不会自动跳转")
        
        # 显示位置在"分类选择后"的自定义列
        display_custom_columns_by_position('分类选择后', df, current_idx)
        
        # 备注区域
        st.markdown("### 📝 备注")
        
        # 获取或初始化当前备注
        note_key = f"note_{current_idx}"
        if note_key not in st.session_state.notes:
            existing_note = df.iloc[current_idx].get('备注', '') if '备注' in df.columns else ''
            if pd.isna(existing_note):
                existing_note = ''
            st.session_state.notes[note_key] = existing_note
        
        # 备注输入框
        current_note = st.text_area(
            "在此输入备注内容",
            value=st.session_state.notes[note_key],
            height=100,
            key="note_textarea",
            placeholder="输入备注内容...",
            help="备注内容将保存到Excel文件的'备注'列中",
            label_visibility="collapsed"
        )
        
        # 保存当前备注到session
        st.session_state.notes[note_key] = current_note
        st.session_state.current_note = current_note
        
        # 底部导航
        st.markdown("---")
        st.markdown("### 导航控制")
        
        col_bottom1, col_bottom2, col_bottom3 = st.columns([1, 2, 1])
        
        with col_bottom1:
            if st.button("◀ 上一篇", key="bottom_prev", disabled=current_idx <= 0, use_container_width=True):
                # 保存当前备注
                st.session_state.notes[note_key] = current_note
                st.session_state.current_index -= 1
                st.rerun()
        
        with col_bottom2:
            st.markdown(f"**当前文献**: {current_idx + 1} / {len(df)}", help="当前文献序号/总文献数")
        
        with col_bottom3:
            if st.button("下一篇 ▶", key="bottom_next", disabled=current_idx >= len(df) - 1, use_container_width=True):
                # 保存当前备注
                st.session_state.notes[note_key] = current_note
                st.session_state.current_index += 1
                st.rerun()
    
    else:
        # 欢迎界面
        st.info("👈 请在左侧边栏上传Excel文件开始使用")
        
        # 使用说明
        with st.expander("📖 使用说明", expanded=True):
            st.markdown("""
            ### 欢迎使用文献筛选工具！
            
            **主要功能：**
            1. **智能列名识别**：自动检测标题、摘要等字段
            2. **手动列映射**：支持自定义列名对应关系
            3. **自定义列显示**：可选择额外列并设置显示名称、位置和折叠状态
            4. **逐篇筛选**：一次只显示一篇文献，专注阅读
            5. **字体大小调节**：可单独调整摘要和翻译的字体大小
            6. **三种分类**：纳入、排除、待定
            7. **自动跳转**：选择分类后自动跳转到下一篇（可关闭）
            8. **备注功能**：为每篇文献添加个性化备注
            9. **数据导出**：导出处理后的Excel文件
            
            **增强功能：**
            - **自定义列显示**：可以选择数据表中的任意列显示，并配置：
              - 显示名称：为列设置自定义名称
              - 显示位置：选择列显示在哪个区域（原文信息栏、翻译信息栏、分类选择后）
              - 折叠状态：选择是否在折叠区域内显示
            - **多工作表导出**：生成的Excel文件包含四个工作表：
              - 1️⃣ **所有文献**：包含所有文献，用颜色标记分类状态（绿色=纳入，黄色=待定，红色=排除）
              - 2️⃣ **纳入文章**：仅包含标记为"纳入"的文献
              - 3️⃣ **待定文章**：仅包含标记为"待定"的文献
              - 4️⃣ **排除文章**：仅包含标记为"排除"的文献
            
            **使用步骤：**
            1. **上传Excel文件**（左侧边栏）
            2. **配置列映射**（系统会自动检测，您也可以手动调整）
            3. **配置自定义列**（选择要显示的额外列，并设置显示名称、位置和折叠状态）
            4. **调整字体大小**（在左侧边栏的"字体大小设置"中）
            5. **开始筛选**：
               - 阅读文献内容（内容完全展开显示）
               - 点击上方分类按钮进行标记
               - 选择后自动跳转到下一篇（默认开启）
               - 在下方添加备注（可选）
            6. **保存结果**：
               - 完成后点击"保存进度并导出"
               - 下载处理后的Excel文件（包含四个工作表）
            
            **导出效果：**
            - **所有文献**工作表：
              - 保留所有原始数据
              - 添加"备注"列保存您的笔记
              - **纳入**的文献：序号单元格标记为绿色
              - **排除**的文献：序号单元格标记为红色
              - **待定**的文献：序号单元格标记为黄色
            - **分类工作表**：
              - 分别包含对应分类的文献
              - 便于后续整理和分析
            """)
        
        # 示例文件格式
        st.markdown("### 📋 示例Excel格式")
        example_data = {
            '序号': [1, 2, 3],
            '标题': ['人工智能在医学诊断中的应用', '深度学习算法优化研究', '自然语言处理技术进展'],
            '标题翻译': ['Application of AI in Medical Diagnosis', 'Research on Deep Learning Algorithm Optimization', 'Advances in Natural Language Processing Technology'],
            '摘要': ['这篇论文探讨了AI在医疗领域的应用...', '本研究提出了一种新的深度学习优化方法...', '本文综述了近年来NLP技术的发展...'],
            '摘要翻译': ['This paper explores the application of AI in the medical field...', 'This study proposes a new deep learning optimization method...', 'This article reviews the development of NLP technology in recent years...'],
            '作者': ['张三, 李四', '王五, 赵六', '钱七, 孙八'],
            '年份': [2023, 2022, 2021],
            '期刊': ['计算机学报', '软件学报', '中文信息学报'],
            '关键词': ['人工智能, 医疗诊断', '深度学习, 优化算法', '自然语言处理, 综述'],
            '备注': ['重要参考文献', '方法新颖', '综述文章']
        }
        st.dataframe(pd.DataFrame(example_data), use_container_width=True)

def display_custom_columns_by_position(position, df, current_idx):
    """按位置显示自定义列"""
    if not st.session_state.extra_columns:
        return
    
    # 获取该位置的所有列
    cols_in_position = []
    for col_name, col_config in st.session_state.extra_columns.items():
        if col_config['position'] == position:
            cols_in_position.append((col_name, col_config))
    
    if not cols_in_position:
        return
    
    # 按折叠状态分组
    direct_cols = []  # 不折叠的列
    collapsed_cols = []  # 折叠的列
    
    for col_name, col_config in cols_in_position:
        if col_config['collapsed']:
            collapsed_cols.append((col_name, col_config))
        else:
            direct_cols.append((col_name, col_config))
    
    # 显示不折叠的列
    for col_name, col_config in direct_cols:
        if col_name in df.columns:
            value = df.iloc[current_idx][col_name]
            if pd.notna(value):
                st.markdown(f"**{col_config['display_name']}**")
                display_custom_column_value(value, col_name, current_idx)
    
    # 显示折叠的列
    if collapsed_cols:
        with st.expander("📋 更多信息", expanded=False):
            for col_name, col_config in collapsed_cols:
                if col_name in df.columns:
                    value = df.iloc[current_idx][col_name]
                    if pd.notna(value):
                        st.markdown(f"**{col_config['display_name']}**")
                        display_custom_column_value(value, col_name, current_idx)

# ====================== 运行应用 ======================
if __name__ == "__main__":
    main()